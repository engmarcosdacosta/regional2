import json
import re
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
KMZ_PATH = Path(r"C:\Users\engma\Documents\EFC - Copia.kmz")
SOURCE_KMZ_PATH = Path(r"C:\Users\engma\Documents\EFC - Copia.kmz.bak")
AXIS_KMZ_PATH = Path(r"C:\Users\engma\Documents\EFC.kmz")
XLSX_PATH = Path(r"C:\Users\engma\Documents\EFC KM A KM DETALHADO.xlsx")
OUTPUT_PATH = ROOT / "public" / "data" / "efc-data.json"

KML_NS = {"kml": "http://www.opengis.net/kml/2.2"}
SEGMENT_RANGES = {
    "EFC I": (0, 314),
    "EFC II": (315, 599),
    "EFC III": (600, 9999),
}
SEGMENT_COLORS = {
    "EFC I": "#1ea672",
    "EFC II": "#ef8c22",
    "EFC III": "#6b3df0",
}
CATEGORY_META = {
    "HOTBOX": {"label": "HotBox", "color": "#f15c4c"},
    "INTEGRIDADE": {"label": "Integridade", "color": "#f0b429"},
    "PRAD": {"label": "PRAD", "color": "#2f9e44"},
    "PONTILHAO": {"label": "Pontilhão", "color": "#1c7ed6"},
    "OBRA": {"label": "Obra", "color": "#495057"},
}


def normalize(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).upper().strip()


def category_for(title: str) -> str:
    normalized = normalize(title)
    for key in ("HOTBOX", "INTEGRIDADE", "PRAD", "PONTILHAO"):
        if normalized.startswith(key):
            return key
    return "OBRA"


def segment_for_km(km: int) -> str:
    for segment, (start, end) in SEGMENT_RANGES.items():
        if start <= km <= end:
            return segment
    return "EFC III"


def parse_coord(text: str) -> list[float]:
    lon, lat, *_ = [float(value) for value in text.split(",")]
    return [lat, lon]


def load_lines() -> list[dict]:
    with zipfile.ZipFile(KMZ_PATH, "r") as archive:
        root = ET.fromstring(archive.read("doc.kml"))

    lines = []
    for document in root.findall(".//kml:Document", KML_NS):
        name_el = document.find("kml:name", KML_NS)
        coordinates_el = document.find(".//kml:LineString/kml:coordinates", KML_NS)
        if name_el is None or coordinates_el is None or not coordinates_el.text:
            continue

        points = [
            parse_coord(row.strip())
            for row in coordinates_el.text.splitlines()
            if row.strip()
        ]
        if not points:
            continue

        lines.append(
            {
                "id": normalize(name_el.text).replace(" ", "_"),
                "name": name_el.text,
                "color": SEGMENT_COLORS.get(name_el.text, "#495057"),
                "points": points,
            }
        )

    return lines


def load_km_coordinates() -> dict[int, list[float]]:
    with zipfile.ZipFile(SOURCE_KMZ_PATH, "r") as archive:
        root = ET.fromstring(archive.read("doc.kml"))

    km_coords: dict[int, list[float]] = {}
    for placemark in root.findall(".//kml:Placemark", KML_NS):
        description = placemark.find("kml:description", KML_NS)
        coordinates_el = placemark.find(".//kml:Point/kml:coordinates", KML_NS)
        if description is None or coordinates_el is None or not description.text:
            continue
        match = re.search(r"ESTACA</B>\s*=\s*(\d+)", description.text)
        if not match:
            continue
        km = int(match.group(1))
        km_coords.setdefault(km, parse_coord(coordinates_el.text.strip()))

    return km_coords


def load_axis_km_coordinates() -> dict[int, list[float]]:
    with zipfile.ZipFile(AXIS_KMZ_PATH, "r") as archive:
        root = ET.fromstring(archive.read("doc.kml"))

    km_coords: dict[int, list[float]] = {}
    for placemark in root.findall(".//kml:Placemark", KML_NS):
        description = placemark.find("kml:description", KML_NS)
        coordinates_el = placemark.find(".//kml:Point/kml:coordinates", KML_NS)
        if description is None or coordinates_el is None or not description.text:
            continue
        match = re.search(r"ESTACA</B>\s*=\s*(\d+)", description.text)
        if not match:
            continue
        km = int(match.group(1))
        km_coords.setdefault(km, parse_coord(coordinates_el.text.strip()))

    return km_coords


def load_works(km_coords: dict[int, list[float]]) -> list[dict]:
    workbook = load_workbook(XLSX_PATH, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    works = []
    item_id = 1
    for row in worksheet.iter_rows(values_only=True):
        title = "" if row[0] is None else str(row[0]).strip()
        detail = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        kms = []
        for match in re.finditer(r"KM\s*-?\s*(\d+)", title, flags=re.IGNORECASE):
            km = int(match.group(1))
            if km not in kms:
                kms.append(km)

        category = category_for(title)
        category_meta = CATEGORY_META[category]
        for km in kms:
            if km not in km_coords:
                continue
            works.append(
                {
                    "id": f"obra-{item_id}",
                    "km": km,
                    "segment": segment_for_km(km),
                    "category": category,
                    "categoryLabel": category_meta["label"],
                    "color": category_meta["color"],
                    "title": title,
                    "detail": detail,
                    "position": km_coords[km],
                }
            )
            item_id += 1

    works.sort(key=lambda item: (item["km"], item["category"], item["title"]))
    return works


def load_km_milestones(axis_km_coords: dict[int, list[float]]) -> list[dict]:
    milestones = []
    for km, position in sorted(axis_km_coords.items()):
        if km % 10 != 0:
            continue
        milestones.append(
            {
                "id": f"km-{km}",
                "km": km,
                "segment": segment_for_km(km),
                "position": position,
            }
        )
    return milestones


def build_summary(works: list[dict]) -> dict:
    categories = {}
    segments = {}
    kms = [work["km"] for work in works]

    for key, meta in CATEGORY_META.items():
        count = sum(1 for work in works if work["category"] == key)
        if count:
            categories[key] = {**meta, "count": count}

    for segment in SEGMENT_RANGES:
        count = sum(1 for work in works if work["segment"] == segment)
        segments[segment] = {"count": count, "color": SEGMENT_COLORS[segment]}

    return {
        "totalWorks": len(works),
        "coveredKms": len(set(kms)),
        "kmRange": {"min": min(kms), "max": max(kms)},
        "categories": categories,
        "segments": segments,
    }


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    km_coords = load_km_coordinates()
    axis_km_coords = load_axis_km_coordinates()
    works = load_works(km_coords)
    km_milestones = load_km_milestones(axis_km_coords)
    data = {
        "generatedAt": datetime.now().isoformat(),
        "lines": load_lines(),
        "works": works,
        "kmMilestones": km_milestones,
        "summary": build_summary(works),
    }
    OUTPUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Arquivo gerado em {OUTPUT_PATH}")
    print(f"Linhas: {len(data['lines'])} | Obras: {len(works)} | Marcos KM: {len(km_milestones)}")


if __name__ == "__main__":
    main()
